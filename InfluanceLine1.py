import os
import math
import openpyxl
#import panda as pd
from openpyxl import load_workbook

wb = load_workbook(filename = "InfluenceLine.xlsx")
#wb = openpyxl.Workbook()
sheet = wb["IRC"]

########################################################################################
# Written by: Radha Krishna Mallik, PhD Scholor, Institute of Engineering Pulchowk ampus#
#########################################################################################



#### READ IRC_A LOADING DATA FROM MASTER EXCEL FILE "InfluenceLine.xlsx"#######
###############################################################################
def IrcALoadSpacing():
    P = []
    for i in range (2,10):
        P1 = 2*sheet.cell(row = 7,column = 11-i).value
        P.append (P1)
    S = []
    for i in range (2,9):
        S1 = sheet.cell(row = 10,column = 10-i).value
        S.append (S1)
    return [P,S]
###############################################################################
#### READ IRC_70R LOADING DATA FROM MASTER EXCEL FILE "InfluenceLine.xlsx"#######
###############################################################################
def Irc70RLoadSpacing():
    P = []
    for i in range (2,10):
        P1 = sheet.cell(row = 22,column = 11-i).value
        P.append (P1)
    S = []
    for i in range (2,9):
        S1 = sheet.cell(row = 25,column = 10-i).value
        S.append (S1)
    return [P,S]

###############################################################################
ircAspacing = IrcALoadSpacing()[1]
ircALoad = IrcALoadSpacing()[0]
irc70Rspacing = Irc70RLoadSpacing()[1]
irc70RLoad = Irc70RLoadSpacing()[0]
##############################################################################

class AbsMaxMoment:
    def __init__(self,span,CW,NG,GS,fck,Ig,N,SN,load,spacing):
        self.span = span
        self.CW = CW       ## CarriageWay Width of the Bridge
        self.NG = NG       ## No of Girder in the Bridge
        self.GS =GS        ## Girder to Girder Spacing
        self.N = N         ##No of steps for vehicle movement in whole girder
        self.SN = SN       ##No of sections in girder for output
        self.load = load   ## IRC CLass A Loading or IRC_70R Loading
        self.spacing = spacing ## Spacing between Axles of Vehicle
        sheet["B1"].value = span ### Bridge Span
        self.IF = 1+(4.5/(6+span))## Impact Factor
        self.fck =fck
        self.Ig = Ig
        self.E = 5000000*self.fck**(0.5)   #kN/m^2
        sheet["B2"].value = NG
        sheet["B4"].value = CW
        sheet["B3"].value = self.IF
        sheet["E33"].value = GS                  
    def loadNposFromRight(self,headPosition):
        posFL = [0]
        x0 = 0
        for i in range(0,7):
            x0 = x0 + self.spacing[i]
            posFL.append(x0)        
        posFromRight = []

        PL = []
        for i in range(0,8):        
            P = self.IF*self.load[i]        
            d1 = headPosition+posFL[i]
            if d1>=self.span:
                P = 0
            PL.append(P)
            posFromRight.append(d1)
        return [posFromRight,PL]

    def momentAtSectionXfromLeft(self,load,sectionPosLeft,posFromRight):
        posFromLeft = self.span-posFromRight
        RA = load*posFromRight/self.span
        RB = load*posFromLeft/self.span
        if posFromLeft < sectionPosLeft:
            M = RB*(self.span - sectionPosLeft)
            V = RB
            d = 1000*load*posFromRight*sectionPosLeft*(self.span**2-sectionPosLeft**2-posFromRight**2)/(6*self.span*self.E*self.Ig*self.NG)
        else:
            M = RA*sectionPosLeft
            V = RA
            d = 1000*load*(self.span*(sectionPosLeft-posFromLeft)**3-(self.span**2-posFromRight**2)*sectionPosLeft*posFromRight-posFromRight*sectionPosLeft**3)/(6*self.span*self.E*self.Ig*self.NG)
        return [M,V,d]

    def MomentNshear(self):
        MomentInfluence = []
        ShearInfluence = []
        DeflectionInfluence = []
        absMaxMoment = []
        absMaxShear = []
        absMaxDeflection = []
        sectionList = []
        vehicleHead = []
        for i in range(0,self.SN+1):
            sectionPosLeft = i*self.span/self.SN
            sectionList.append(sectionPosLeft)
            MomentInfluence = []
            ShearInfluence = []
            DeflectionInfluence = []
            for j in range(0,self.N+1):
                vehicleHeadFromRight = j*self.span/self.N       
                posFromRight=self.loadNposFromRight(vehicleHeadFromRight)[0]
                updatedLoad = self.loadNposFromRight(vehicleHeadFromRight)[1]
                vehicleHead.append(vehicleHeadFromRight)
                Moment = 0
                Shear = 0
                Deflection = 0
                for i in range(0,7):
                    p = posFromRight[i]
                    l = updatedLoad[i]
                    Moment = Moment+self.momentAtSectionXfromLeft(l,sectionPosLeft,p)[0]
                    Shear = Shear + self.momentAtSectionXfromLeft(l,sectionPosLeft,p)[1]
                    Deflection = Deflection + self.momentAtSectionXfromLeft(l,sectionPosLeft,p)[2]
                MomentInfluence.append(Moment)
                ShearInfluence.append(Shear)
                DeflectionInfluence.append(Deflection)
            absMax = max(MomentInfluence)
            absMaxV = max(ShearInfluence)
            absMaxD = max(DeflectionInfluence)
            absMaxMoment.append(absMax)
            absMaxShear.append(absMaxV)
            absMaxDeflection.append(absMaxD)
            MomentInfluence = []
            ShearInfluence = []
        return [sectionList,absMaxMoment,absMaxShear,absMaxDeflection]


##############################################################################
#OBJECT FORMATION span,CW,NG,GS,fck,Ig,N,SN,load,spacing
###############################################################################
s = float(input(" Enter Span of the Bridge,m: "))
cw = float(input(" Enter CarriageWayWidth of the Bridge,m: "))
ng = int(input(" Enter no of girder in bridge: "))
gs = float(input("Enter girder to girder spacing in bridge: "))
fck = float(input("Enter grade of concrete,MPa: "))
ig = float(input("Enter Moment of Inertia of each girder,m^4: "))
si = float(input("Enter output section interval(Span/si = integer),m: "))
# OBJECT FOR IRC_A and IRC_70R LOADING
SN = int(s/si)
a = AbsMaxMoment(s,cw,ng,gs,fck,ig,100,SN,ircALoad,ircAspacing)
b = AbsMaxMoment(s,cw,ng,gs,fck,ig,100,SN,irc70RLoad,irc70Rspacing)

print("######################################################################")
print(" Double Lane_IRC_Class A: Max absolute Bending Moment in bridge spline")      
print("######################################################################")      

print(a.MomentNshear()[0])             
print(a.MomentNshear()[1])
print(a.MomentNshear()[2])            
print(a.MomentNshear()[3])
print("######################################################################")
print(" Single Lane IRC_70R: Max absolute Bending Moment in bridge spline")      
print("######################################################################")

print(b.MomentNshear()[0])             
print(b.MomentNshear()[1])
print(b.MomentNshear()[2])
print(b.MomentNshear()[3]) 
###############################################################################
# Writing output in Excel File 2-IRC CLASS A LOADING
###############################################################################
section = a.MomentNshear()[0]             
momentAtSection = a.MomentNshear()[1]
shearAtSection = a.MomentNshear()[2] 
deflectionAtSection = a.MomentNshear()[3]

for i in range(0,SN+1):
   
    sheet.cell(row = 44+i,column = 1).value = section[i]
    
    sheet.cell(row = 44+i,column = 2).value = momentAtSection[i]
    
    sheet.cell(row = 44+i,column = 3).value = shearAtSection[i]
    
    sheet.cell(row = 44+i,column = 8).value = deflectionAtSection[i]
    
###############################################################################
# Writing output in Excel File 1-IRC CLASS 70R LOADING
###############################################################################
section = b.MomentNshear()[0]             
momentAtSection = b.MomentNshear()[1]
shearAtSection = b.MomentNshear()[2] 
deflectionAtSection = b.MomentNshear()[3]

for i in range(0,SN+1):
    
    sheet.cell(row = 74+i,column = 1).value = section[i]
   
    sheet.cell(row = 74+i,column = 2).value = momentAtSection[i]
   
    sheet.cell(row = 74+i,column = 3).value = shearAtSection[i]
    
    sheet.cell(row = 74+i,column = 8).value = deflectionAtSection[i]
    
###############################################################################
# Save as a New Excel File "40M_VehicleOutput.xlsx" 
###############################################################################    
wb.save(str(s)+"M_"+str(ng)+"G_"+str(cw)+"MCW_VehicleOutput.xlsx")




